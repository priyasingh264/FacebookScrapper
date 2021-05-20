# Post to scrap :- https://www.facebook.com/urbanladder/photos/a.947076545302763/4215002548510130/

from selenium import webdriver
import time
from selenium.webdriver.opera.options import Options
import openpyxl as xl


def add_data_to_excel( row_number, column_number, data):
    c = sheet.cell(row=row_number, column=column_number)
    c.value = data
    print(data)


# Opening the excel file and reading the URL
wb = xl.load_workbook('Web Scrapping Task..xlsx')
sheet = wb['Sheet2']
row_value = sheet.cell(row=2, column=2)
print(row_value.value)
url_value = row_value.value
print(url_value)


# Using opera driver for automation
# Running in headless mode
opt = Options()
opt.add_argument("--headless")
driver = webdriver.Opera(
    executable_path='C:\\Users\\hp\\Downloads\\Driver For '
                    'Automation\\operadriver_win64\\operadriver_win64\\operadriver.exe')

# Opening the site in full screen mode
driver.maximize_window()
driver.get(url_value)

# Wait for Page to Load
time.sleep(30)

# Login to facebook
driver.find_element_by_xpath("(//*[@class='m9osqain jq4qci2q a3bd9o3v'])[1]").click()
driver.find_element_by_xpath("(//*[@class='m9osqain jq4qci2q a3bd9o3v'])[1]").send_keys("ENTER YOUR FB ID")
driver.find_element_by_xpath("(//*[@class='m9osqain jq4qci2q a3bd9o3v'])[2]").click()
driver.find_element_by_xpath("(//*[@class='m9osqain jq4qci2q a3bd9o3v'])[2]").send_keys("ENETR YOUR PASSWORD")
driver.find_element_by_xpath("(//*[@class='a8c37x1j ni8dbmo4 stjgntxs l9j0dhe7 ltmttdrg g0qnabr5'])[2]").click()
time.sleep(30)

# Adding current url in row 4 column 1
add_data_to_excel(4, 1, driver.current_url)


# Adding name in row 4 column 2
name = driver.find_element_by_xpath("(//*[@class='nc684nl6']/a)[2]").text
add_data_to_excel(4, 2, name)


# Adding number of likes in row 4 column 3
likes = driver.find_element_by_xpath("(//*[@class='pcp91wgn'])").text
add_data_to_excel(4, 3, likes)

driver.find_element_by_xpath("(//*[@class='pcp91wgn'])").click()
time.sleep(3)

# Name of people who liked the post
count_names = len(driver.find_elements_by_xpath("(//*[@class='j83agx80 cbu4d94t ew0dbk1b irj2b8pg']//*["
                                                "@class='q9uorilb'])"))
print(count_names)

list_of_names = " "
url_of_id = " "

for i in range(1, count_names ):
    checkAllName = "(//*[@class='j83agx80 cbu4d94t ew0dbk1b irj2b8pg']//*[@class='q9uorilb'])" + "[" + str(i) + "]/a";
    list_of_names = list_of_names + ', ' + driver.find_element_by_xpath(checkAllName).text
    url_of_id = url_of_id + ", " + driver.find_element_by_xpath(checkAllName).get_attribute("href")
add_data_to_excel(4, 4, list_of_names)
add_data_to_excel(4, 5, url_of_id)


# Closing the name popup
driver.find_element_by_xpath("(//*[@class='cypi58rs pmk7jnqg fcg2cn6m tkr6xdv7'])").click()

# Number of people who commented the post
number_of_comment = driver.find_element_by_xpath("(//*[@class='gtad4xkn'])[1]//span").text
add_data_to_excel(4, 9, number_of_comment )


# Name of people who commented the post
list_of_names = " "
url_of_id = " "
count_names = len(driver.find_elements_by_xpath("(//*[@class='nc684nl6'])"))
print(count_names)

i = 3
while i < count_names - 1:
    checkAllName = "(//*[@class='nc684nl6'])" + "[" + str(i) + "]/a"
    list_of_names = list_of_names + ', ' + driver.find_element_by_xpath(checkAllName).text + " "
    i = i + 1
    checkAllName = "(//*[@class='nc684nl6'])" + "[" + str(i) + "]/a"
    url_of_id = url_of_id + ", " + driver.find_element_by_xpath(checkAllName).get_attribute("href") + " "

add_data_to_excel(4, 10, list_of_names)
add_data_to_excel(4, 11, url_of_id)

wb.save("Output_file_2.xlsx")
driver.close()
